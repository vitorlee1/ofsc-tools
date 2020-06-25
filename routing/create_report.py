import logging, json, pprint
import argparse
import time,csv
import openpyxl
from openpyxl import Workbook, load_workbook, utils
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import (
    AreaChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

def init_script():
    #TODO Add option to start from an intermediate step
    # Parse arguments
    global args
    parser = argparse.ArgumentParser()
    parser.add_argument("--verbose", type=int, help="Verbosity Level 0,1,2,3 (default is 1)",  choices = { 0, 1, 2, 3}, default = 1)
    parser.add_argument("--output", type=str, help="output file to be used (default is report.xlsx) ", default = "report.xlsx")
    parser.add_argument("--routing_input_file",  type=str, help="CSV files to be processed. Default is routing_output.csv", default = "routing_output.csv")
    parser.add_argument("-template", type=str, help="Excel template to use. Default is routing_template.xlsx", default = "routing_template.xlsx")
    args = parser.parse_args()

    # create logger
    global logger
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    logger.info("Starting execution: {}".format(args.verbose))
    # create console handler and set level to debug
    ch = logging.StreamHandler()
    # create formatter
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    # add formatter to ch
    ch.setFormatter(formatter)
    logger.addHandler(ch)
    logger.setLevel(40 - 10*int(args.verbose))
    logger.info("Log level is {}".format(args.verbose))



init_script()
wb = load_workbook(args.template)
ws = wb["routing_data"]
table = ws.tables["Routing_Groups"]
logger.info("Table found in {}".format(table.ref))
boundaries = openpyxl.utils.cell.range_boundaries(table.ref)
logger.debug(boundaries)
initial_col = boundaries[0]
initial_row = boundaries[1]
logger.info("Initial Row: {}, Initial Column: {}".format(boundaries[0], boundaries[1]))

# Read CSV

with open(args.routing_input_file, newline='') as csvfile:
    reader = csv.reader(csvfile, delimiter=',')
    i=initial_row
    final_col = 0
    for row in reader:
            logger.debug(row)
            if final_col ==0:
                final_col = len(row)
            for j,cell in enumerate(row):
                if i>initial_row:
                    logger.debug ("Inserting {},{} - {}".format(i,j+initial_col,cell))
                    ws.cell(row=i, column=j+initial_col, value = int(cell))
                else:
                    ws.cell(row=i, column=j+initial_col, value = cell)
                    logger.info("Adding {} cols".format(len(row)))
            i+=1

    final_row = i - 1
logger.info("Final Row: {}, Final Col: {}".format(final_row, final_col))
table_range = "{}{}:{}{}".format(openpyxl.utils.cell.get_column_letter(initial_col), initial_row, openpyxl.utils.cell.get_column_letter(final_col), final_row)
logger.info("Setting table range to {}".format(table_range))
table.ref=table_range

# Add Routing StackedLine Chart

chart = AreaChart()
data = Reference(ws, min_col=2, min_row=1, max_col=final_col, max_row=final_row)
cats = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=final_row)
chart.add_data(data, titles_from_data=True)
chart.grouping = "percentStacked"
chart.set_categories(cats)
ws.add_chart(chart, "H1")

wb.save(filename=args.output)









